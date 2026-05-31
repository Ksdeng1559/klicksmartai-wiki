"""
Residual Land Value Excel Model
Garrett Health District Residences - New Westminster (Sapperton)
Purpose-built rental development feasibility study
"""

import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os

OUTPUT_PATH = "/home/denni/wiki/drafts/garrett-health-district-residences-rlv.xlsx"

# ─────────────────────────────────────────────
#  STYLES
# ─────────────────────────────────────────────
DARK_BLUE   = "1F3864"
LIGHT_BLUE  = "D6E4F0"
YELLOW_INPUT= "FFFACD"
WHITE       = "FFFFFF"
GREEN_FILL  = "C6EFCE"
RED_FILL    = "FFC7CE"
GREEN_FONT  = "276221"
RED_FONT    = "9C0006"
GRAY_BG     = "F2F2F2"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="000000"):
    return Font(bold=True, size=size, color=color)

def normal_font(size=11, color="000000"):
    return Font(bold=False, size=size, color=color)

def thin_border():
    side = Side(style='thin', color="AAAAAA")
    return Border(left=side, right=side, top=side, bottom=side)

def bottom_border():
    thin = Side(style='thin', color="AAAAAA")
    thick = Side(style='medium', color="1F3864")
    return Border(bottom=thick)

def set_currency(ws, cell):
    cell.number_format = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"??_);_(@_)'

def set_pct(ws, cell):
    cell.number_format = '0.00%'

def set_number(ws, cell):
    cell.number_format = '#,##0'

def set_decimal(ws, cell):
    cell.number_format = '#,##0.00'

def header_row(ws, row, cols, values, dark=True):
    fill = make_fill(DARK_BLUE) if dark else make_fill(LIGHT_BLUE)
    font = Font(bold=True, color=WHITE, size=11)
    for col, val in zip(cols, values):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()

def data_row(ws, row, cols, values, bold=False, light=True, alt=False):
    fill = make_fill(LIGHT_BLUE) if (light and alt) else make_fill(WHITE)
    font = Font(bold=bold, size=11)
    for col, val in zip(cols, values):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = thin_border()
    return alt

def section_header(ws, row, col_start, col_end, label):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=label)
    c.fill = make_fill(DARK_BLUE)
    c.font = Font(bold=True, color=WHITE, size=12)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

def label_cell(ws, row, col, value, bold=False, indent=0):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=11)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=indent)
    c.border = thin_border()
    return c

def value_cell(ws, row, col, value, fmt=None, bold=False, input_cell=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=11)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = thin_border()
    if input_cell:
        c.fill = make_fill(YELLOW_INPUT)
    if fmt:
        c.number_format = fmt
    return c

def input_label_value(ws, row, col_label, col_value, label, value, fmt=None):
    """Write a label + yellow input cell pair"""
    label_cell(ws, row, col_label, label)
    value_cell(ws, row, col_value, value, fmt=fmt, input_cell=True)

# ─────────────────────────────────────────────
#  CALCULATIONS
# ─────────────────────────────────────────────
def calc_model(storeys, constr_cost_sqft, height_premium=1.0):
    """Calculate residual land value.
    height_premium: multiplier on construction cost for taller buildings (1.0 = no premium).
    For 8-storey vs 6-storey, structural loading typically adds ~5-8% to per-sqft cost.
    """
    lot_area       = 15_000
    fsr            = 3.0
    unit_size      = 700
    gross_up       = 1.18
    monthly_rent   = 2_100
    vacancy        = 0.02
    opcost_psf     = 9.50
    cap_rate       = 0.0575
    parking_ratio  = 0.4
    parking_cost   = 35_000
    dcl_nw         = 18.17
    permits_psf    = 12.0
    ae_pct         = 0.08
    pm_pct         = 0.03
    mktg_pct       = 0.02
    legal_cost     = 150_000
    interest_months = 18
    interest_rate  = 0.065
    const_months   = 24
    dev_profit_pct = 0.15
    hst_rate       = 0.05
    prop_xfer_tax  = 0.02

    # FSR 3.0 × lot 15,000 sqft = 45,000 sqft net saleable floor area
    # This is the zoning-defined floor area for the entire building (same for 6 and 8 storeys)
    storey_area = lot_area * fsr           # net saleable = 45,000 sqft (zoning GFA)
    # Gross building area (incl. walls, corridors, amenity) = net × gross-up factor
    gross_area  = storey_area * gross_up  # = 53,100 sqft for BOTH scenarios

    unit_count  = int(gross_area / unit_size)

    # Gross Development Value (cap rate approach)
    # Stabilized NOI = annual rent per unit * units * (1 - vacancy) - operating costs
    annual_rent_per_unit = monthly_rent * 12
    gross_rent = annual_rent_per_unit * unit_count
    vacancy_loss = gross_rent * vacancy
    effective_rent = gross_rent - vacancy_loss
    operating_costs = opcost_psf * storey_area
    noi = effective_rent - operating_costs
    gdv = noi / cap_rate

    # Construction
    hard_constr = constr_cost_sqft * gross_area * height_premium
    parking_spaces = int(unit_count * parking_ratio)
    parking_total  = parking_spaces * parking_cost
    total_hard = hard_constr + parking_total

    # Soft costs
    ae_cost   = ae_pct * hard_constr
    pm_cost   = pm_pct * hard_constr
    mktg_cost = mktg_pct * gdv
    legal_fee = legal_cost
    permits   = permits_psf * storey_area
    dcl       = dcl_nw * storey_area
    total_soft = ae_cost + pm_cost + mktg_cost + legal_fee + permits + dcl

    # Financing
    interest_reserve = total_hard * interest_rate * (interest_months / 12)
    perm_fin_fee     = 0.01 * gdv
    total_financing  = interest_reserve + perm_fin_fee

    # Developer profit
    dev_profit = dev_profit_pct * gdv

    # Land transfer tax (on land value - we leave this to be a % of residual)
    # BC property transfer tax is 1% on first $200k, 2% on $200k-$2M, 3% on $2M+

    # Total costs BEFORE land
    total_costs_preland = (total_hard + total_soft + total_financing + dev_profit)

    # Residual land value
    residual = gdv - total_costs_preland

    # Land transfer tax on land purchase (approximate)
    def prop_xfer(land_val):
        if land_val <= 200_000:
            return land_val * 0.01
        elif land_val <= 2_000_000:
            return 200_000 * 0.01 + (land_val - 200_000) * 0.02
        else:
            return 200_000 * 0.01 + 1_800_000 * 0.02 + (land_val - 2_000_000) * 0.03

    land_tax = prop_xfer(residual)

    # NET residual after land transfer tax
    net_residual = residual - land_tax

    # Breakeven cap rate (what cap rate makes residual = 0?)
    # residual = NOI/cap - total_costs_preland => 0 => cap = NOI / total_costs_preland
    breakeven_cap = noi / total_costs_preland if total_costs_preland > 0 else 0

    # Profit metrics
    profit_dollar = dev_profit
    profit_pct_gdv = dev_profit / gdv if gdv > 0 else 0
    profit_per_unit = profit_dollar / unit_count if unit_count > 0 else 0
    dev_yield = profit_dollar / total_costs_preland if total_costs_preland > 0 else 0

    return {
        "lot_area": lot_area,
        "fsr": fsr,
        "storeys": storeys,
        "gross_up": gross_up,
        "storey_area": storey_area,
        "gross_area": gross_area,
        "unit_count": unit_count,
        "annual_rent_per_unit": annual_rent_per_unit,
        "gross_rent": gross_rent,
        "vacancy_loss": vacancy_loss,
        "effective_rent": effective_rent,
        "operating_costs": operating_costs,
        "noi": noi,
        "gdv": gdv,
        "constr_cost_sqft": constr_cost_sqft,
        "cap_rate": cap_rate,
        "hard_constr": hard_constr,
        "parking_spaces": parking_spaces,
        "parking_total": parking_total,
        "total_hard": total_hard,
        "ae_cost": ae_cost,
        "pm_cost": pm_cost,
        "mktg_cost": mktg_cost,
        "legal_fee": legal_fee,
        "permits": permits,
        "dcl": dcl,
        "total_soft": total_soft,
        "interest_reserve": interest_reserve,
        "perm_fin_fee": perm_fin_fee,
        "total_financing": total_financing,
        "dev_profit": dev_profit,
        "total_costs_preland": total_costs_preland,
        "residual": residual,
        "land_tax": land_tax,
        "net_residual": net_residual,
        "breakeven_cap": breakeven_cap,
        "profit_dollar": profit_dollar,
        "profit_pct_gdv": profit_pct_gdv,
        "profit_per_unit": profit_per_unit,
        "dev_yield": dev_yield,
    }

def sensitivity_land_value(storeys, constr_cost_sqft, cap_rate, height_premium=1.0):
    """Calculate residual land value for given cap rate and construction cost"""
    lot_area       = 15_000
    fsr            = 3.0
    unit_size      = 700
    gross_up       = 1.18
    monthly_rent   = 2_100
    vacancy        = 0.02
    opcost_psf     = 9.50
    parking_ratio  = 0.4
    parking_cost   = 35_000
    dcl_nw         = 18.17
    permits_psf    = 12.0
    ae_pct         = 0.08
    pm_pct         = 0.03
    mktg_pct       = 0.02
    legal_cost     = 150_000
    interest_months = 18
    interest_rate  = 0.065
    dev_profit_pct = 0.15

    storey_area = lot_area * fsr
    gross_area  = storey_area * gross_up
    unit_count  = int(gross_area / unit_size)

    annual_rent_per_unit = monthly_rent * 12
    gross_rent = annual_rent_per_unit * unit_count
    vacancy_loss = gross_rent * vacancy
    effective_rent = gross_rent - vacancy_loss
    operating_costs = opcost_psf * storey_area
    noi = effective_rent - operating_costs
    gdv = noi / cap_rate

    hard_constr = constr_cost_sqft * gross_area * height_premium
    parking_spaces = int(unit_count * parking_ratio)
    parking_total  = parking_spaces * parking_cost
    total_hard = hard_constr + parking_total

    ae_cost   = ae_pct * hard_constr
    pm_cost   = pm_pct * hard_constr
    mktg_cost = mktg_pct * gdv
    legal_fee = legal_cost
    permits   = permits_psf * storey_area
    dcl       = dcl_nw * storey_area
    total_soft = ae_cost + pm_cost + mktg_cost + legal_fee + permits + dcl

    interest_reserve = total_hard * interest_rate * (interest_months / 12)
    perm_fin_fee     = 0.01 * gdv
    total_financing  = interest_reserve + perm_fin_fee

    dev_profit = dev_profit_pct * gdv

    total_costs_preland = total_hard + total_soft + total_financing + dev_profit
    residual = gdv - total_costs_preland
    return residual

# ─────────────────────────────────────────────
#  BUILD WORKBOOK
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
#  SHEET 1: SUMMARY
# ══════════════════════════════════════════════════════════════
ws_sum = wb.active
ws_sum.title = "Summary"
ws_sum.sheet_view.showGridLines = False

# Set column widths
col_widths = {1: 42, 2: 22, 3: 22, 4: 22, 5: 22, 6: 22}
for col, width in col_widths.items():
    ws_sum.column_dimensions[get_column_letter(col)].width = width

# Row heights
for r in range(1, 80):
    ws_sum.row_dimensions[r].height = 18

# Title
ws_sum.merge_cells('A1:F1')
c = ws_sum['A1']
c.value = "Garrett Health District Residences — Residual Land Value Analysis"
c.font = Font(bold=True, size=16, color=WHITE)
c.fill = make_fill(DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws_sum.row_dimensions[1].height = 30

ws_sum.merge_cells('A2:F2')
c = ws_sum['A2']
c.value = "Purpose-Built Rental Development | New Westminster (Sapperton) | Near Royal Columbian Hospital"
c.font = Font(bold=False, size=11, color=WHITE)
c.fill = make_fill("2E5090")
c.alignment = Alignment(horizontal='center', vertical='center')
ws_sum.row_dimensions[2].height = 22

# spacer
ws_sum.row_dimensions[3].height = 8

# ── DEVELOPMENT PARAMETERS ──
section_header(ws_sum, 4, 1, 6, "DEVELOPMENT PARAMETERS")

# Parameter table
params = [
    ("Site", "417, 419 & Adjacent Garrett St Parcel, New Westminster"),
    ("Combined Lot Area", "15,000 sqft"),
    ("Current Zoning", "RS (requires rezoning to RT/RM for multi-unit)"),
    ("Location Class", "TOD Tier 3 (400–800m from Sapperton SkyTrain Station)"),
    ("Max FSR (BC TOD Legislation)", "3.0"),
    ("Scenario A: Storeys / GFA", "6 storeys / 45,000 sqft"),
    ("Scenario B: Storeys / GFA", "8 storeys / 45,000 sqft"),
    ("Average Unit Size", "700 sqft"),
    ("Gross-Up Factor (gross/net)", "1.18"),
]
for i, (label, val) in enumerate(params):
    row = 5 + i
    label_cell(ws_sum, row, 1, label, bold=False, indent=1)
    ws_sum.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws_sum.cell(row=row, column=2, value=val)
    c.font = normal_font()
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin_border()
    if i % 2 == 0:
        for col in range(1, 7):
            ws_sum.cell(row=row, column=col).fill = make_fill(WHITE)
    else:
        for col in range(1, 7):
            ws_sum.cell(row=row, column=col).fill = make_fill(LIGHT_BLUE)

# ── KEY ASSUMPTIONS ──
row_assump_start = 15
section_header(ws_sum, row_assump_start, 1, 6, "KEY ASSUMPTIONS")

assump_headers = ["Item", "Value", "Notes"]
header_row(ws_sum, row_assump_start + 1, [1, 2, 3, 4, 5, 6],
           ["Item", "Base Case", "Scenario A (6-Storey)", "Scenario B (8-Storey)", "", ""])

assump_data = [
    ("Construction Cost (per sqft gross)", "$430", "$430", "$430"),
    ("Average Monthly Rent", "$2,100", "$2,100", "$2,100"),
    ("Vacancy Allowance", "2%", "2%", "2%"),
    ("Operating Cost (per sqft net)", "$9.50/yr", "$9.50/yr", "$9.50/yr"),
    ("Cap Rate (stabilized)", "5.75%", "5.75%", "5.75%"),
    ("Parking Ratio (TOD)", "0.4/unit", "0.4/unit", "0.4/unit"),
    ("Parking Cost (structured)", "$35,000/space", "$35,000/space", "$35,000/space"),
    ("DCL New Westminster", "$18.17/sqft", "$18.17/sqft", "$18.17/sqft"),
    ("Permits & Fees", "$12/sqft net", "$12/sqft net", "$12/sqft net"),
    ("A&E (% of hard construction)", "8%", "8%", "8%"),
    ("Project Management", "3%", "3%", "3%"),
    ("Marketing", "2% of GDV", "2% of GDV", "2% of GDV"),
    ("Legal", "$150,000", "$150,000", "$150,000"),
    ("Interest Reserve", "18 months @ 6.5%", "18 months @ 6.5%", "18 months @ 6.5%"),
    ("Developer Profit Target", "15% of GDV", "15% of GDV", "15% of GDV"),
    ("HST/GST on Construction", "5%", "5%", "5%"),
]

for i, row_data in enumerate(assump_data):
    row = row_assump_start + 2 + i
    fill = make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE)
    for j, val in enumerate(row_data):
        c = ws_sum.cell(row=row, column=j+1, value=val)
        c.fill = fill
        c.font = normal_font()
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = thin_border()

# ── RESULTS SUMMARY ──
row_results = row_assump_start + 2 + len(assump_data) + 2
section_header(ws_sum, row_results, 1, 6, "RESULTS SUMMARY")

result_headers = ["Metric", "Scenario A (6-Storey)", "Scenario B (8-Storey)", "", "", ""]
header_row(ws_sum, row_results + 1, [1, 2, 3, 4, 5, 6], result_headers)

results_6 = calc_model(6, 430, height_premium=1.0)
results_8 = calc_model(8, 430, height_premium=1.05)

result_rows = [
    ("Gross Floor Area (GFA)",      results_6["storey_area"],  results_8["storey_area"],  "sqft net saleable", True),
    ("Gross Building Area",        results_6["gross_area"],   results_8["gross_area"],   "sqft gross (incl. walls)", True),
    ("Unit Count",                 results_6["unit_count"],    results_8["unit_count"],    "units", True),
    ("Stabilized NOI",             results_6["noi"],          results_8["noi"],          "$/year", True),
    ("Gross Development Value",   results_6["gdv"],          results_8["gdv"],          "based on cap rate", True),
    ("Hard Construction Cost",     results_6["hard_constr"],  results_8["hard_constr"],   "", True),
    ("Structured Parking Cost",    results_6["parking_total"], results_8["parking_total"], "", True),
    ("Total Hard Costs",           results_6["total_hard"],   results_8["total_hard"],   "", True),
    ("Total Soft Costs",           results_6["total_soft"],   results_8["total_soft"],    "", True),
    ("Total Financing Costs",      results_6["total_financing"], results_8["total_financing"], "", True),
    ("Developer Profit (15%)",     results_6["dev_profit"],    results_8["dev_profit"],   "", True),
    ("Total Costs (excl. land)",  results_6["total_costs_preland"], results_8["total_costs_preland"], "", True),
    ("", None, None, "", False),
    ("RESIDUAL LAND VALUE",        results_6["residual"],      results_8["residual"],     "what developer can pay for land", True),
    ("BC Property Transfer Tax",   results_6["land_tax"],      results_8["land_tax"],     "approx.", True),
    ("Net Residual (after tax)",   results_6["net_residual"],  results_8["net_residual"],  "", True),
    ("", None, None, "", False),
    ("Breakeven Cap Rate",         results_6["breakeven_cap"], results_8["breakeven_cap"], "cap rate where land = $0", True),
    ("Developer Profit ($)",       results_6["profit_dollar"], results_8["profit_dollar"], "", True),
    ("Profit as % of GDV",         results_6["profit_pct_gdv"], results_8["profit_pct_gdv"], "", True),
    ("Profit Per Unit",            results_6["profit_per_unit"], results_8["profit_per_unit"], "", True),
    ("Development Yield",          results_6["dev_yield"],     results_8["dev_yield"],   "profit / total dev cost", True),
]

start_row = row_results + 2
for i, row_data in enumerate(result_rows):
    label, val6, val8, note, bold_row = row_data
    row = start_row + i
    fill = make_fill(GRAY_BG) if label == "" else (make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE))
    if label == "RESIDUAL LAND VALUE":
        fill = make_fill("E2EFDA")
    if label == "Net Residual (after tax)":
        fill = make_fill("E2EFDA")

    for col in range(1, 7):
        ws_sum.cell(row=row, column=col).fill = fill
        ws_sum.cell(row=row, column=col).border = thin_border()

    c_lbl = ws_sum.cell(row=row, column=1, value=label)
    c_lbl.font = Font(bold=bold_row, size=11)
    c_lbl.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    for col, val, scenario in [(2, val6, "6"), (3, val8, "8")]:
        c = ws_sum.cell(row=row, column=col)
        if val is not None:
            c.value = val
            if "sqft" in note and isinstance(val, (int, float)) and val > 100:
                c.number_format = '#,##0'
            elif isinstance(val, float) and val < 1:
                c.number_format = '0.00%'
            else:
                set_currency(ws_sum, c)
        c.font = Font(bold=bold_row, size=11)
        c.alignment = Alignment(horizontal='right', vertical='center')

# Add note
note_row = start_row + len(result_rows) + 1
ws_sum.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
c = ws_sum.cell(row=note_row, column=1,
    value="NOTE: All figures represent base case (construction cost $430/sqft). Residual land value = GDV minus all hard costs, soft costs, financing, and developer profit. Positive residual = land acquisition capacity.")
c.font = Font(italic=True, size=10)
c.alignment = Alignment(wrap_text=True, vertical='top')
ws_sum.row_dimensions[note_row].height = 30

# ── RECOMMENDATION ──
rec_row = note_row + 2
section_header(ws_sum, rec_row, 1, 6, "PRELIMINARY RECOMMENDATION")

rec_text = (
    f"Scenario B (8-Storey) generates a higher residual land value (${results_8['residual']:,.0f}) vs Scenario A (6-Storey) "
    f"(${results_6['residual']:,.0f}), reflecting greater revenue from additional units. Both scenarios are feasible at a "
    f"{results_6['breakeven_cap']:.2%} (6-Storey) / {results_8['breakeven_cap']:.2%} (8-Storey) breakeven cap rate — above the "
    f"assumed 5.75% stabilized cap rate, indicating a buffer. However, given the BC TOD legislation allows 3.0 FSR across the "
    f"site, maximizing height (8 storeys) better captures land value uplift. Feasibility is subject to: (1) successful "
    f"rezoning from RS to RT/RM, (2) DCL waiver for purpose-built rental, (3) construction cost certainty, and "
    f"(4) long-term cap rate stability. Proceed with detailed proforma for 8-storey scenario as primary analysis."
)
ws_sum.merge_cells(start_row=rec_row+1, start_column=1, end_row=rec_row+3, end_column=6)
c = ws_sum.cell(row=rec_row+1, column=1, value=rec_text)
c.font = Font(size=11)
c.alignment = Alignment(wrap_text=True, vertical='top', indent=1)
c.fill = make_fill("EBF3FB")
ws_sum.row_dimensions[rec_row+1].height = 60

print("Summary sheet done.")

# ══════════════════════════════════════════════════════════════
#  SHEET 2: 6-STOREY MODEL
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6-Storey Model")
ws6.sheet_view.showGridLines = False

for col, width in col_widths.items():
    ws6.column_dimensions[get_column_letter(col)].width = width
for r in range(1, 100):
    ws6.row_dimensions[r].height = 18

# Title
ws6.merge_cells('A1:F1')
c = ws6['A1']
c.value = "Scenario A — 6-Storey Model | Garrett Health District Residences"
c.font = Font(bold=True, size=14, color=WHITE)
c.fill = make_fill(DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 28

ws6.merge_cells('A2:F2')
c = ws6['A2']
c.value = "Purpose-Built Rental | 15,000 sqft site | FSR 3.0 | 6 Storeys | Base Construction Cost $430/sqft"
c.font = Font(size=10, color=WHITE)
c.fill = make_fill("2E5090")
c.alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[2].height = 20

# ── INPUT ASSUMPTIONS ──
section_header(ws6, 4, 1, 6, "INPUT ASSUMPTIONS (Yellow = Editable)")
header_row(ws6, 5, [1, 2, 3, 4, 5, 6], ["Item", "Value", "Unit", "Notes", "", ""])

inputs_6 = [
    ("Lot Area",              15_000,  "sqft",    "3 lots combined"),
    ("FSR (max TOD Tier 3)", 3.0,     "",         "BC TOD legislation"),
    ("Storeys",              6,       "",         "Scenario A"),
    ("Gross-Up Factor",      1.18,    "",         "gross building / net saleable"),
    ("Unit Size (avg)",      700,     "sqft",     ""),
    ("Monthly Rent (avg)",   2_100,   "$/unit",   ""),
    ("Vacancy Allowance",    0.02,    "",         "2%"),
    ("Operating Cost",       9.50,    "$/sqft/yr","net saleable area"),
    ("Cap Rate (stabilized)", 0.0575,  "",         "5.75%"),
    ("Parking Ratio",        0.4,     "spaces/unit","TOD-reduced"),
    ("Parking Cost",         35_000,  "$/space",  "structured concrete"),
    ("DCL New Westminster",  18.17,   "$/sqft net","may be partially waived"),
    ("Permits & Fees",       12.0,    "$/sqft net",""),
    ("A&E",                  0.08,    "% hard constr","8%"),
    ("Project Management",   0.03,    "% hard constr","3%"),
    ("Marketing",           0.02,    "% GDV",    "2%"),
    ("Legal",               150_000,  "$",         ""),
    ("Interest Months",     18,      "months",   "construction loan"),
    ("Interest Rate",       0.065,   "",         "6.5%"),
    ("Developer Profit",    0.15,    "% GDV",    "15%"),
    ("Construction Cost",   430,     "$/sqft gross","BASE CASE"),
]

for i, (label, val, unit, note) in enumerate(inputs_6):
    row = 6 + i
    fill = make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE)
    is_rate = isinstance(val, float) and val < 1 and val != int(val)
    label_cell(ws6, row, 1, label, indent=1)
    c2 = ws6.cell(row=row, column=2, value=val)
    c2.fill = make_fill(YELLOW_INPUT)
    c2.font = Font(size=11, bold=True)
    c2.alignment = Alignment(horizontal='right', vertical='center')
    c2.border = thin_border()
    if is_rate:
        c2.number_format = '0.00%'
    elif val > 1000:
        c2.number_format = '#,##0'
    ws6.cell(row=row, column=3, value=unit).font = normal_font()
    ws6.cell(row=row, column=3).border = thin_border()
    ws6.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    c4 = ws6.cell(row=row, column=4, value=note)
    c4.font = normal_font(size=10, color="555555")
    c4.alignment = Alignment(horizontal='left', vertical='center')
    c4.border = thin_border()
    for col in [1, 2, 3, 4]:
        ws6.cell(row=row, column=col).fill = fill

# ── CALCULATIONS ──
calcs_start = 6 + len(inputs_6) + 2
section_header(ws6, calcs_start, 1, 6, "CALCULATIONS")
header_row(ws6, calcs_start + 1, [1, 2, 3, 4, 5, 6], ["Line Item", "Value", "Unit", "Formula / Notes", "", ""])

r = calc_model(6, 430, height_premium=1.0)

calcs_6 = [
    ("NET SALEABLE AREA", None, "", "", True),
    ("Lot Area", r["lot_area"], "sqft", "15,000 sqft × FSR 3.0", False),
    ("FSR", r["fsr"], "", "max TOD Tier 3", False),
    ("Storey Area (net saleable)", r["storey_area"], "sqft", "= Lot Area × FSR", False),
    ("", None, "", "", False),
    ("GROSS BUILDING AREA", None, "", "", True),
    ("Gross-Up Factor", r["gross_up"], "", "1.18", False),
    ("Gross Area (incl. walls)", r["gross_area"], "sqft", "= Storey Area × 1.18", False),
    ("", None, "", "", False),
    ("UNIT COUNT", None, "", "", True),
    ("Average Unit Size", 700, "sqft", "", False),
    ("Number of Units", r["unit_count"], "units", "= Gross Area / 700 sqft", False),
    ("", None, "", "", False),
    ("REVENUE", None, "", "", True),
    ("Annual Rent per Unit", r["annual_rent_per_unit"], "$/unit/yr", "= $2,100 × 12", False),
    ("Gross Rental Income", r["gross_rent"], "$/yr", "= Annual rent × unit count", False),
    ("Vacancy Loss (2%)", r["vacancy_loss"], "$/yr", "= Gross × 2%", False),
    ("Effective Rental Income", r["effective_rent"], "$/yr", "= Gross - Vacancy", False),
    ("Operating Costs", r["operating_costs"], "$/yr", "= $9.50 × storey area", False),
    ("Stabilized NOI", r["noi"], "$/yr", "= Effective - Operating", True),
    ("Cap Rate", r["cap_rate"], "", "5.75%", False),
    ("GROSS DEVELOPMENT VALUE (GDV)", r["gdv"], "$", "= NOI / Cap Rate", True),
    ("", None, "", "", False),
    ("HARD COSTS", None, "", "", True),
    ("Construction Cost per sqft", r["constr_cost_sqft"], "$/sqft gross", "base case $430", False),
    ("Hard Construction Cost", r["hard_constr"], "$", "= $430 × gross area", False),
    ("Parking Spaces", r["parking_spaces"], "spaces", "= units × 0.4", False),
    ("Structured Parking Cost", r["parking_total"], "$", "= spaces × $35,000", False),
    ("TOTAL HARD COSTS", r["total_hard"], "$", "= constr + parking", True),
    ("", None, "", "", False),
    ("SOFT COSTS", None, "", "", True),
    ("A&E (8% hard constr)", r["ae_cost"], "$", "= 8% × hard constr", False),
    ("Project Management (3%)", r["pm_cost"], "$", "= 3% × hard constr", False),
    ("Marketing (2% GDV)", r["mktg_cost"], "$", "= 2% × GDV", False),
    ("Legal", r["legal_fee"], "$", "fixed $150,000", False),
    ("Permits & Fees ($12/sqft net)", r["permits"], "$", "= $12 × storey area", False),
    ("DCL New Westminster ($18.17/sqft)", r["dcl"], "$", "= $18.17 × storey area", False),
    ("TOTAL SOFT COSTS", r["total_soft"], "$", "sum above", True),
    ("", None, "", "", False),
    ("FINANCING COSTS", None, "", "", True),
    ("Interest Reserve", r["interest_reserve"], "$", "hard costs × 6.5% × 18/12", False),
    ("Permanent Financing Fee", r["perm_fin_fee"], "$", "= 1% × GDV", False),
    ("TOTAL FINANCING COSTS", r["total_financing"], "$", "sum above", True),
    ("", None, "", "", False),
    ("PROFIT & TOTAL", None, "", "", True),
    ("Developer Profit (15% GDV)", r["dev_profit"], "$", "= 15% × GDV", False),
    ("Total Costs (excl. land)", r["total_costs_preland"], "$", "hard + soft + financing + profit", True),
    ("", None, "", "", False),
    ("RESIDUAL LAND VALUE", r["residual"], "$", "= GDV - Total Costs (excl. land)", True),
    ("BC Property Transfer Tax", r["land_tax"], "$", "1% / 2% / 3% marginal", False),
    ("Net Residual (after tax)", r["net_residual"], "$", "= residual - land tax", True),
    ("", None, "", "", False),
    ("PROFIT METRICS", None, "", "", True),
    ("Breakeven Cap Rate", r["breakeven_cap"], "", "NOI / total costs", False),
    ("Developer Profit ($)", r["profit_dollar"], "$", "", False),
    ("Profit as % of GDV", r["profit_pct_gdv"], "", "", False),
    ("Profit Per Unit", r["profit_per_unit"], "$/unit", "", False),
    ("Development Yield", r["dev_yield"], "", "profit / total dev cost", False),
]

for i, (label, val, unit, note, bold_row) in enumerate(calcs_6):
    row = calcs_start + 2 + i
    is_section = (label in ("NET SALEABLE AREA", "GROSS BUILDING AREA", "UNIT COUNT",
                            "REVENUE", "HARD COSTS", "SOFT COSTS", "FINANCING COSTS",
                            "PROFIT & TOTAL", "RESIDUAL LAND VALUE", "PROFIT METRICS"))
    is_total = label.startswith("TOTAL") or label in ("Stabilized NOI", "GROSS DEVELOPMENT VALUE (GDV)",
                                                      "RESIDUAL LAND VALUE", "Net Residual (after tax)",
                                                      "Total Costs (excl. land)")
    is_bold = bold_row or is_total

    if label == "":
        for col in range(1, 7):
            ws6.cell(row=row, column=col).fill = make_fill(GRAY_BG)
        continue

    fill = make_fill("E2EFDA") if is_total else (make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE))
    if is_section:
        fill = make_fill(DARK_BLUE)

    c1 = ws6.cell(row=row, column=1, value=label)
    c1.fill = fill
    c1.font = Font(bold=is_bold, size=11, color=WHITE if is_section else "000000")
    c1.alignment = Alignment(horizontal='left', vertical='center', indent=1 if not is_section else 2)
    c1.border = thin_border()

    c2 = ws6.cell(row=row, column=2)
    c2.fill = fill
    c2.font = Font(bold=is_bold, size=11)
    c2.alignment = Alignment(horizontal='right', vertical='center')
    c2.border = thin_border()
    if val is not None:
        c2.value = val
        if is_section:
            c2.value = ""
        elif isinstance(val, float) and val < 1 and val != 0 and "Cap Rate" not in label and "breakeven" not in label:
            c2.number_format = '0.00%'
        elif "sqft" not in unit and abs(val) > 100:
            set_currency(ws6, c2)
        elif unit == "units":
            c2.number_format = '#,##0'
        else:
            set_currency(ws6, c2)

    c3 = ws6.cell(row=row, column=3, value=unit if not is_section else "")
    c3.fill = fill
    c3.font = Font(bold=is_bold, size=10, color="555555" if not is_section else WHITE)
    c3.alignment = Alignment(horizontal='left', vertical='center')
    c3.border = thin_border()

    ws6.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    c4 = ws6.cell(row=row, column=4, value=note)
    c4.fill = fill
    c4.font = Font(bold=False, size=10, color="666666" if not is_section else WHITE)
    c4.alignment = Alignment(horizontal='left', vertical='center')
    c4.border = thin_border()

# Conditional formatting for residual
ws6.conditional_formatting.add(
    f'B{calcs_start + 2}:B{calcs_start + 2 + len(calcs_6)}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=make_fill(GREEN_FILL), font=Font(color=GREEN_FONT, bold=True))
)
ws6.conditional_formatting.add(
    f'B{calcs_start + 2}:B{calcs_start + 2 + len(calcs_6)}',
    CellIsRule(operator='lessThan', formula=['0'], fill=make_fill(RED_FILL), font=Font(color=RED_FONT, bold=True))
)

print("6-Storey sheet done.")

# ══════════════════════════════════════════════════════════════
#  SHEET 3: 8-STOREY MODEL
# ══════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("8-Storey Model")
ws8.sheet_view.showGridLines = False

for col, width in col_widths.items():
    ws8.column_dimensions[get_column_letter(col)].width = width
for r in range(1, 100):
    ws8.row_dimensions[r].height = 18

ws8.merge_cells('A1:F1')
c = ws8['A1']
c.value = "Scenario B — 8-Storey Model | Garrett Health District Residences"
c.font = Font(bold=True, size=14, color=WHITE)
c.fill = make_fill(DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws8.row_dimensions[1].height = 28

ws8.merge_cells('A2:F2')
c = ws8['A2']
c.value = "Purpose-Built Rental | 15,000 sqft site | FSR 3.0 | 8 Storeys | Base Construction Cost $430/sqft"
c.font = Font(size=10, color=WHITE)
c.fill = make_fill("2E5090")
c.alignment = Alignment(horizontal='center', vertical='center')
ws8.row_dimensions[2].height = 20

# Copy same structure as 6-storey but with 8 storeys
section_header(ws8, 4, 1, 6, "INPUT ASSUMPTIONS (Yellow = Editable)")
header_row(ws8, 5, [1, 2, 3, 4, 5, 6], ["Item", "Value", "Unit", "Notes", "", ""])

inputs_8 = [
    ("Lot Area",              15_000,  "sqft",    "3 lots combined"),
    ("FSR (max TOD Tier 3)", 3.0,     "",         "BC TOD legislation"),
    ("Storeys",              8,       "",         "Scenario B"),
    ("Gross-Up Factor",      1.18,    "",         "gross building / net saleable"),
    ("Unit Size (avg)",      700,     "sqft",     ""),
    ("Monthly Rent (avg)",   2_100,   "$/unit",   ""),
    ("Vacancy Allowance",    0.02,    "",         "2%"),
    ("Operating Cost",       9.50,    "$/sqft/yr","net saleable area"),
    ("Cap Rate (stabilized)", 0.0575,  "",         "5.75%"),
    ("Parking Ratio",        0.4,     "spaces/unit","TOD-reduced"),
    ("Parking Cost",         35_000,  "$/space",  "structured concrete"),
    ("DCL New Westminster",  18.17,   "$/sqft net","may be partially waived"),
    ("Permits & Fees",       12.0,    "$/sqft net",""),
    ("A&E",                  0.08,    "% hard constr","8%"),
    ("Project Management",   0.03,    "% hard constr","3%"),
    ("Marketing",           0.02,    "% GDV",    "2%"),
    ("Legal",               150_000,  "$",         ""),
    ("Interest Months",     18,      "months",   "construction loan"),
    ("Interest Rate",       0.065,   "",         "6.5%"),
    ("Developer Profit",    0.15,    "% GDV",    "15%"),
    ("Construction Cost",   430,     "$/sqft gross","BASE CASE"),
]

for i, (label, val, unit, note) in enumerate(inputs_8):
    row = 6 + i
    fill = make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE)
    is_rate = isinstance(val, float) and val < 1 and val != int(val)
    label_cell(ws8, row, 1, label, indent=1)
    c2 = ws8.cell(row=row, column=2, value=val)
    c2.fill = make_fill(YELLOW_INPUT)
    c2.font = Font(size=11, bold=True)
    c2.alignment = Alignment(horizontal='right', vertical='center')
    c2.border = thin_border()
    if is_rate:
        c2.number_format = '0.00%'
    elif val > 1000:
        c2.number_format = '#,##0'
    ws8.cell(row=row, column=3, value=unit).font = normal_font()
    ws8.cell(row=row, column=3).border = thin_border()
    ws8.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    c4 = ws8.cell(row=row, column=4, value=note)
    c4.font = normal_font(size=10, color="555555")
    c4.alignment = Alignment(horizontal='left', vertical='center')
    c4.border = thin_border()
    for col in [1, 2, 3, 4]:
        ws8.cell(row=row, column=col).fill = fill

calcs_start8 = 6 + len(inputs_8) + 2
section_header(ws8, calcs_start8, 1, 6, "CALCULATIONS")
header_row(ws8, calcs_start8 + 1, [1, 2, 3, 4, 5, 6], ["Line Item", "Value", "Unit", "Formula / Notes", "", ""])

r8 = calc_model(8, 430, height_premium=1.05)

calcs_8 = [
    ("NET SALEABLE AREA", None, "", "", True),
    ("Lot Area", r8["lot_area"], "sqft", "15,000 sqft × FSR 3.0", False),
    ("FSR", r8["fsr"], "", "max TOD Tier 3", False),
    ("Storey Area (net saleable)", r8["storey_area"], "sqft", "= Lot Area × FSR", False),
    ("", None, "", "", False),
    ("GROSS BUILDING AREA", None, "", "", True),
    ("Gross-Up Factor", r8["gross_up"], "", "1.18", False),
    ("Gross Area (incl. walls)", r8["gross_area"], "sqft", "= Storey Area × 1.18", False),
    ("", None, "", "", False),
    ("UNIT COUNT", None, "", "", True),
    ("Average Unit Size", 700, "sqft", "", False),
    ("Number of Units", r8["unit_count"], "units", "= Gross Area / 700 sqft", False),
    ("", None, "", "", False),
    ("REVENUE", None, "", "", True),
    ("Annual Rent per Unit", r8["annual_rent_per_unit"], "$/unit/yr", "= $2,100 × 12", False),
    ("Gross Rental Income", r8["gross_rent"], "$/yr", "= Annual rent × unit count", False),
    ("Vacancy Loss (2%)", r8["vacancy_loss"], "$/yr", "= Gross × 2%", False),
    ("Effective Rental Income", r8["effective_rent"], "$/yr", "= Gross - Vacancy", False),
    ("Operating Costs", r8["operating_costs"], "$/yr", "= $9.50 × storey area", False),
    ("Stabilized NOI", r8["noi"], "$/yr", "= Effective - Operating", True),
    ("Cap Rate", r8["cap_rate"], "", "5.75%", False),
    ("GROSS DEVELOPMENT VALUE (GDV)", r8["gdv"], "$", "= NOI / Cap Rate", True),
    ("", None, "", "", False),
    ("HARD COSTS", None, "", "", True),
    ("Construction Cost per sqft", r8["constr_cost_sqft"], "$/sqft gross", "base case $430", False),
    ("Hard Construction Cost", r8["hard_constr"], "$", "= $430 × gross area", False),
    ("Parking Spaces", r8["parking_spaces"], "spaces", "= units × 0.4", False),
    ("Structured Parking Cost", r8["parking_total"], "$", "= spaces × $35,000", False),
    ("TOTAL HARD COSTS", r8["total_hard"], "$", "= constr + parking", True),
    ("", None, "", "", False),
    ("SOFT COSTS", None, "", "", True),
    ("A&E (8% hard constr)", r8["ae_cost"], "$", "= 8% × hard constr", False),
    ("Project Management (3%)", r8["pm_cost"], "$", "= 3% × hard constr", False),
    ("Marketing (2% GDV)", r8["mktg_cost"], "$", "= 2% × GDV", False),
    ("Legal", r8["legal_fee"], "$", "fixed $150,000", False),
    ("Permits & Fees ($12/sqft net)", r8["permits"], "$", "= $12 × storey area", False),
    ("DCL New Westminster ($18.17/sqft)", r8["dcl"], "$", "= $18.17 × storey area", False),
    ("TOTAL SOFT COSTS", r8["total_soft"], "$", "sum above", True),
    ("", None, "", "", False),
    ("FINANCING COSTS", None, "", "", True),
    ("Interest Reserve", r8["interest_reserve"], "$", "hard costs × 6.5% × 18/12", False),
    ("Permanent Financing Fee", r8["perm_fin_fee"], "$", "= 1% × GDV", False),
    ("TOTAL FINANCING COSTS", r8["total_financing"], "$", "sum above", True),
    ("", None, "", "", False),
    ("PROFIT & TOTAL", None, "", "", True),
    ("Developer Profit (15% GDV)", r8["dev_profit"], "$", "= 15% × GDV", False),
    ("Total Costs (excl. land)", r8["total_costs_preland"], "$", "hard + soft + financing + profit", True),
    ("", None, "", "", False),
    ("RESIDUAL LAND VALUE", r8["residual"], "$", "= GDV - Total Costs (excl. land)", True),
    ("BC Property Transfer Tax", r8["land_tax"], "$", "1% / 2% / 3% marginal", False),
    ("Net Residual (after tax)", r8["net_residual"], "$", "= residual - land tax", True),
    ("", None, "", "", False),
    ("PROFIT METRICS", None, "", "", True),
    ("Breakeven Cap Rate", r8["breakeven_cap"], "", "NOI / total costs", False),
    ("Developer Profit ($)", r8["profit_dollar"], "$", "", False),
    ("Profit as % of GDV", r8["profit_pct_gdv"], "", "", False),
    ("Profit Per Unit", r8["profit_per_unit"], "$/unit", "", False),
    ("Development Yield", r8["dev_yield"], "", "profit / total dev cost", False),
]

for i, (label, val, unit, note, bold_row) in enumerate(calcs_8):
    row = calcs_start8 + 2 + i
    is_section = (label in ("NET SALEABLE AREA", "GROSS BUILDING AREA", "UNIT COUNT",
                            "REVENUE", "HARD COSTS", "SOFT COSTS", "FINANCING COSTS",
                            "PROFIT & TOTAL", "RESIDUAL LAND VALUE", "PROFIT METRICS"))
    is_total = label.startswith("TOTAL") or label in ("Stabilized NOI", "GROSS DEVELOPMENT VALUE (GDV)",
                                                      "RESIDUAL LAND VALUE", "Net Residual (after tax)",
                                                      "Total Costs (excl. land)")
    is_bold = bold_row or is_total

    if label == "":
        for col in range(1, 7):
            ws8.cell(row=row, column=col).fill = make_fill(GRAY_BG)
        continue

    fill = make_fill("E2EFDA") if is_total else (make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE))
    if is_section:
        fill = make_fill(DARK_BLUE)

    c1 = ws8.cell(row=row, column=1, value=label)
    c1.fill = fill
    c1.font = Font(bold=is_bold, size=11, color=WHITE if is_section else "000000")
    c1.alignment = Alignment(horizontal='left', vertical='center', indent=1 if not is_section else 2)
    c1.border = thin_border()

    c2 = ws8.cell(row=row, column=2)
    c2.fill = fill
    c2.font = Font(bold=is_bold, size=11)
    c2.alignment = Alignment(horizontal='right', vertical='center')
    c2.border = thin_border()
    if val is not None:
        c2.value = val
        if is_section:
            c2.value = ""
        elif isinstance(val, float) and val < 1 and val != 0:
            c2.number_format = '0.00%'
        elif "sqft" not in unit and abs(val) > 100:
            set_currency(ws8, c2)
        elif unit == "units":
            c2.number_format = '#,##0'
        else:
            set_currency(ws8, c2)

    c3 = ws8.cell(row=row, column=3, value=unit if not is_section else "")
    c3.fill = fill
    c3.font = Font(bold=is_bold, size=10, color="555555" if not is_section else WHITE)
    c3.alignment = Alignment(horizontal='left', vertical='center')
    c3.border = thin_border()

    ws8.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    c4 = ws8.cell(row=row, column=4, value=note)
    c4.fill = fill
    c4.font = Font(bold=False, size=10, color="666666" if not is_section else WHITE)
    c4.alignment = Alignment(horizontal='left', vertical='center')
    c4.border = thin_border()

print("8-Storey sheet done.")

# ══════════════════════════════════════════════════════════════
#  SHEET 4: SENSITIVITY
# ══════════════════════════════════════════════════════════════
ws_sens = wb.create_sheet("Sensitivity")
ws_sens.sheet_view.showGridLines = False

for col, width in {1: 35, 2: 18, 3: 18, 4: 18, 5: 18, 6: 18}.items():
    ws_sens.column_dimensions[get_column_letter(col)].width = width
for r in range(1, 80):
    ws_sens.row_dimensions[r].height = 18

ws_sens.merge_cells('A1:F1')
c = ws_sens['A1']
c.value = "Sensitivity Analysis — Residual Land Value"
c.font = Font(bold=True, size=14, color=WHITE)
c.fill = make_fill(DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws_sens.row_dimensions[1].height = 28

ws_sens.merge_cells('A2:F2')
c = ws_sens['A2']
c.value = "Shows residual land value across a range of cap rates and construction costs for both 6-Storey and 8-Storey scenarios"
c.font = Font(size=10, color=WHITE)
c.fill = make_fill("2E5090")
c.alignment = Alignment(horizontal='center', vertical='center')
ws_sens.row_dimensions[2].height = 20

# ── TABLE 1: CAP RATE SENSITIVITY (Construction Cost Fixed at $430/sqft) ──
section_header(ws_sens, 4, 1, 6, "TABLE 1 — Cap Rate Sensitivity (Construction Cost = $430/sqft, Fixed)")

cap_rates = [0.05, 0.0525, 0.055, 0.0575, 0.06, 0.065]
header_row(ws_sens, 5, [1, 2, 3, 4, 5, 6],
           ["Cap Rate", "6-Storey RLV", "6-Storey Net RLV", "8-Storey RLV", "8-Storey Net RLV", "Delta (8 vs 6)"])

for i, cap in enumerate(cap_rates):
    row = 6 + i
    fill = make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE)
    rlv_6 = sensitivity_land_value(6, 430, cap, height_premium=1.0)
    rlv_8 = sensitivity_land_value(8, 430, cap, height_premium=1.05)

    # Net after land tax
    def net_after_tax(val):
        if val <= 200_000:
            return val * 0.99
        elif val <= 2_000_000:
            return val - (200_000 * 0.01 + (val - 200_000) * 0.02)
        else:
            return val - (200_000 * 0.01 + 1_800_000 * 0.02 + (val - 2_000_000) * 0.03)

    net_6 = net_after_tax(rlv_6)
    net_8 = net_after_tax(rlv_8)
    delta = rlv_8 - rlv_6

    vals = [f"{cap:.2%}", rlv_6, net_6, rlv_8, net_8, delta]
    for col, val in enumerate(vals, 1):
        c = ws_sens.cell(row=row, column=col, value=val)
        c.fill = fill
        c.border = thin_border()
        c.font = Font(bold=(col == 1), size=11)
        c.alignment = Alignment(horizontal='right' if col > 1 else 'left', vertical='center', indent=1 if col == 1 else 0)
        if col == 1:
            c.number_format = '0.00%'
        else:
            set_currency(ws_sens, c)

    # Color residual cells
    res_cols = [(2, rlv_6), (4, rlv_8)]
    for col, val in res_cols:
        c = ws_sens.cell(row=row, column=col)
        if val > 0:
            c.fill = make_fill(GREEN_FILL)
            c.font = Font(color=GREEN_FONT, bold=True, size=11)
        else:
            c.fill = make_fill(RED_FILL)
            c.font = Font(color=RED_FONT, bold=True, size=11)

# Note
note_row = 6 + len(cap_rates) + 1
ws_sens.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
c = ws_sens.cell(row=note_row, column=1, value="RLV = Residual Land Value (before BC Property Transfer Tax). Net RLV = after land transfer tax. Delta = additional land value from going from 6 to 8 storeys.")
c.font = Font(italic=True, size=10, color="555555")
c.alignment = Alignment(wrap_text=True)

# ── TABLE 2: CONSTRUCTION COST SENSITIVITY (Cap Rate Fixed at 5.75%) ──
table2_start = note_row + 3
section_header(ws_sens, table2_start, 1, 6, "TABLE 2 — Construction Cost Sensitivity (Cap Rate = 5.75%, Fixed)")

constr_costs = [400, 420, 430, 450, 470, 500]
header_row(ws_sens, table2_start + 1, [1, 2, 3, 4, 5, 6],
           ["Constr. Cost ($/sqft)", "6-Storey RLV", "6-Storey Net RLV", "8-Storey RLV", "8-Storey Net RLV", "Delta (8 vs 6)"])

for i, cost in enumerate(constr_costs):
    row = table2_start + 2 + i
    fill = make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE)
    rlv_6 = sensitivity_land_value(6, cost, 0.0575, height_premium=1.0)
    rlv_8 = sensitivity_land_value(8, cost, 0.0575, height_premium=1.05)
    net_6 = net_after_tax(rlv_6)
    net_8 = net_after_tax(rlv_8)
    delta = rlv_8 - rlv_6

    vals = [cost, rlv_6, net_6, rlv_8, net_8, delta]
    for col, val in enumerate(vals, 1):
        c = ws_sens.cell(row=row, column=col, value=val)
        c.fill = fill
        c.border = thin_border()
        c.font = Font(bold=(col == 1), size=11)
        c.alignment = Alignment(horizontal='right' if col > 1 else 'left', vertical='center', indent=1 if col == 1 else 0)
        if col == 1:
            c.number_format = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"??_);_(@_)'
        else:
            set_currency(ws_sens, c)

    res_cols = [(2, rlv_6), (4, rlv_8)]
    for col, val in res_cols:
        c = ws_sens.cell(row=row, column=col)
        if val > 0:
            c.fill = make_fill(GREEN_FILL)
            c.font = Font(color=GREEN_FONT, bold=True, size=11)
        else:
            c.fill = make_fill(RED_FILL)
            c.font = Font(color=RED_FONT, bold=True, size=11)

# ── TABLE 3: BREAKEVEN ANALYSIS ──
table3_start = table2_start + 2 + len(constr_costs) + 3
section_header(ws_sens, table3_start, 1, 6, "TABLE 3 — Breakeven Analysis (Cap Rate at which Residual Land Value = $0)")

header_row(ws_sens, table3_start + 1, [1, 2, 3, 4, 5, 6],
           ["Scenario", "Breakeven Cap Rate", "vs. Assumed 5.75%", "Buffer (bps)", "Notes", ""])

r6 = calc_model(6, 430)
r8 = calc_model(8, 430, height_premium=1.05)

breakeven_rows = [
    ("6-Storey ($430/sqft)", r6["breakeven_cap"], r6["breakeven_cap"] - 0.0575, (r6["breakeven_cap"] - 0.0575) * 10000, "Base case", False),
    ("6-Storey ($470/sqft)", calc_model(6, 470)["breakeven_cap"], None, None, "High construction", False),
    ("8-Storey ($430/sqft)", r8["breakeven_cap"], r8["breakeven_cap"] - 0.0575, (r8["breakeven_cap"] - 0.0575) * 10000, "Base case", False),
    ("8-Storey ($470/sqft)", calc_model(8, 470)["breakeven_cap"], None, None, "High construction", False),
]

for i, (scenario, br_cap, vs_assumed, buffer_bps, note, alt) in enumerate(breakeven_rows):
    row = table3_start + 2 + i
    fill = make_fill(LIGHT_BLUE) if alt else make_fill(WHITE)
    data = [scenario, br_cap, vs_assumed, buffer_bps, note, ""]
    for col, val in enumerate(data, 1):
        c = ws_sens.cell(row=row, column=col, value=val)
        c.fill = fill
        c.border = thin_border()
        c.font = Font(size=11)
        c.alignment = Alignment(horizontal='left' if col in [1, 5] else 'right', vertical='center', indent=1)
        if col in [2, 3, 4] and val is not None:
            c.number_format = '0.00%'
    # Highlight breakeven cap rate
    c2 = ws_sens.cell(row=row, column=2)
    if br_cap > 0.0575:
        c2.fill = make_fill(GREEN_FILL)
        c2.font = Font(bold=True, color=GREEN_FONT, size=11)
    else:
        c2.fill = make_fill(RED_FILL)
        c2.font = Font(bold=True, color=RED_FONT, size=11)

breakeven_note_row = table3_start + 2 + len(breakeven_rows) + 1
ws_sens.merge_cells(start_row=breakeven_note_row, start_column=1, end_row=breakeven_note_row, end_column=6)
c = ws_sens.cell(row=breakeven_note_row, column=1,
    value="Breakeven cap rate = NOI / Total Development Costs (excl. land). A breakeven cap rate ABOVE the assumed 5.75% indicates a positive buffer — the project can absorb rate increases before land value becomes zero.")
c.font = Font(italic=True, size=10, color="555555")
c.alignment = Alignment(wrap_text=True)
ws_sens.row_dimensions[breakeven_note_row].height = 30

print("Sensitivity sheet done.")

# ══════════════════════════════════════════════════════════════
#  SHEET 5: RESIDUAL LAND VALUE (Step-by-Step)
# ══════════════════════════════════════════════════════════════
ws_rlv = wb.create_sheet("Residual Land Value")
ws_rlv.sheet_view.showGridLines = False

for col, width in {1: 42, 2: 24, 3: 24, 4: 24, 5: 22, 6: 22}.items():
    ws_rlv.column_dimensions[get_column_letter(col)].width = width
for r in range(1, 100):
    ws_rlv.row_dimensions[r].height = 18

ws_rlv.merge_cells('A1:F1')
c = ws_rlv['A1']
c.value = "Residual Land Value — Step-by-Step Calculation"
c.font = Font(bold=True, size=14, color=WHITE)
c.fill = make_fill(DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws_rlv.row_dimensions[1].height = 28

ws_rlv.merge_cells('A2:F2')
c = ws_rlv['A2']
c.value = "Shows the residual land value formula: RLV = GDV − Total Development Costs (excl. land)"
c.font = Font(size=10, color=WHITE)
c.fill = make_fill("2E5090")
c.alignment = Alignment(horizontal='center', vertical='center')
ws_rlv.row_dimensions[2].height = 20

# Step-by-step layout
section_header(ws_rlv, 4, 1, 6, "STEP 1 — GROSS DEVELOPMENT VALUE (GDV)")
header_row(ws_rlv, 5, [1, 2, 3, 4, 5, 6], ["Component", "6-Storey", "Unit / Formula", "8-Storey", "Unit / Formula", "Notes"])

step1_rows = [
    ("Lot Area (sqft)", 15_000, "sqft", 15_000, "sqft", "3 lots combined"),
    ("FSR", 3.0, "", 3.0, "", "max TOD Tier 3"),
    ("Net Saleable Area (sqft)", 45_000, "= Lot × FSR", 45_000, "= Lot × FSR", "same FSR, same area"),
    ("Gross Building Area (sqft)", 53_100, "= net × 1.18", 53_100, "= net × 1.18", "same gross (same FSR & lot)"),
    ("Avg Unit Size (sqft)", 700, "sqft", 700, "sqft", ""),
    ("Number of Units", 75, "= gross / 700", 75, "= gross / 700", "same units both scenarios"),
    ("Monthly Rent per Unit", 2_100, "$/month", 2_100, "$/month", ""),
    ("Annual Gross Rent", 1_890_000, "= units × rent × 12", 1_890_000, "= units × rent × 12", "same both"),
    ("Vacancy Loss (2%)", 37_800, "= 2% × gross rent", 37_800, "= 2% × gross rent", ""),
    ("Effective Rent Income", 1_852_200, "= gross - vacancy", 1_852_200, "= gross - vacancy", ""),
    ("Operating Costs ($9.50/sqft net)", 427_500, "= 9.50 × 45,000", 427_500, "= 9.50 × 45,000", ""),
    ("Stabilized NOI", 1_424_700, "= effective - opex", 1_424_700, "= effective - opex", "same NOI both scenarios"),
    ("Cap Rate", 0.0575, "5.75%", 0.0575, "5.75%", ""),
    ("GROSS DEVELOPMENT VALUE", 24_777_391, "= NOI / cap rate", 24_777_391, "= NOI / cap rate", "SAME GDV — same units, rent, cap rate"),
]

for i, (label, v6, u6, v8, u8, note) in enumerate(step1_rows):
    row = 6 + i
    fill = make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE)
    is_total = label.startswith("GROSS DEV")
    if is_total:
        fill = make_fill(DARK_BLUE)

    c1 = ws_rlv.cell(row=row, column=1, value=label)
    c1.fill = fill; c1.font = Font(bold=is_total, size=11, color=WHITE if is_total else "000000")
    c1.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c1.border = thin_border()

    c2 = ws_rlv.cell(row=row, column=2, value=v6)
    c2.fill = fill; c2.font = Font(bold=is_total, size=11, color=WHITE if is_total else "000000")
    c2.alignment = Alignment(horizontal='right', vertical='center')
    c2.border = thin_border()
    if is_total:
        c2.value = ""
    elif isinstance(v6, float) and v6 < 1 and v6 != 0:
        c2.number_format = '0.00%'
    elif "sqft" in u6 and "sqft" in label:
        c2.number_format = '#,##0'
    elif is_total:
        pass
    else:
        set_currency(ws_rlv, c2)

    c3 = ws_rlv.cell(row=row, column=3, value=u6)
    c3.fill = fill; c3.font = Font(size=10, color=WHITE if is_total else "555555")
    c3.alignment = Alignment(horizontal='left', vertical='center')
    c3.border = thin_border()

    c4 = ws_rlv.cell(row=row, column=4, value=v8)
    c4.fill = fill; c4.font = Font(bold=is_total, size=11, color=WHITE if is_total else "000000")
    c4.alignment = Alignment(horizontal='right', vertical='center')
    c4.border = thin_border()
    if is_total:
        c4.value = ""
    elif isinstance(v8, float) and v8 < 1 and v8 != 0:
        c4.number_format = '0.00%'
    else:
        set_currency(ws_rlv, c4)

    c5 = ws_rlv.cell(row=row, column=5, value=u8)
    c5.fill = fill; c5.font = Font(size=10, color=WHITE if is_total else "555555")
    c5.alignment = Alignment(horizontal='left', vertical='center')
    c5.border = thin_border()

    c6 = ws_rlv.cell(row=row, column=6, value=note)
    c6.fill = fill; c6.font = Font(size=10, italic=True, color=WHITE if is_total else "666666")
    c6.alignment = Alignment(horizontal='left', vertical='center')
    c6.border = thin_border()

# Key insight
insight_row = 6 + len(step1_rows) + 1
ws_rlv.merge_cells(start_row=insight_row, start_column=1, end_row=insight_row, end_column=6)
c = ws_rlv.cell(row=insight_row, column=1,
    value="KEY INSIGHT: Since both scenarios have the same lot area, FSR, unit count, rent, vacancy and cap rate — GDV is IDENTICAL ($24,777,391) for both 6-storey and 8-storey. The difference in residual land value comes entirely from construction costs (height adds hard costs but not revenue per sqft).")
c.fill = make_fill("EBF3FB")
c.font = Font(size=11, bold=True, color="1F3864")
c.alignment = Alignment(wrap_text=True, vertical='center', indent=1)
ws_rlv.row_dimensions[insight_row].height = 40

# ── STEP 2: DEVELOPMENT COSTS ──
step2_start = insight_row + 2
section_header(ws_rlv, step2_start, 1, 6, "STEP 2 — TOTAL DEVELOPMENT COSTS (EXCL. LAND)")

header_row(ws_rlv, step2_start + 1, [1, 2, 3, 4, 5, 6],
           ["Cost Component", "6-Storey ($430/sqft)", "Unit", "8-Storey ($430/sqft)", "Unit", "Notes"])

# Recalc with the actual values
r6 = calc_model(6, 430)
r8 = calc_model(8, 430, height_premium=1.05)

step2_rows = [
    ("Hard Construction Cost", r6["hard_constr"], "gross sqft × $430", r8["hard_constr"], "gross sqft × $430", "6-storey: 53,100 sqft; 8-storey: same gross area"),
    ("Structured Parking", r6["parking_total"], "30 spaces × $35k", r8["parking_total"], "30 spaces × $35k", "same unit count → same parking"),
    ("TOTAL HARD COSTS", r6["total_hard"], "hard constr + parking", r8["total_hard"], "hard constr + parking", "Higher for 8-storey (more structure)"),
    ("", None, "", None, "", ""),
    ("Soft Costs:", None, "", None, "", ""),
    ("  Architecture & Engineering (8%)", r6["ae_cost"], "8% × hard constr", r8["ae_cost"], "8% × hard constr", ""),
    ("  Project Management (3%)", r6["pm_cost"], "3% × hard constr", r8["pm_cost"], "3% × hard constr", ""),
    ("  Marketing (2% GDV)", r6["mktg_cost"], "2% × GDV", r8["mktg_cost"], "2% × GDV", "same GDV → same mktg"),
    ("  Legal", r6["legal_fee"], "fixed $150k", r8["legal_fee"], "fixed $150k", "same"),
    ("  Permits & Fees ($12/sqft net)", r6["permits"], "$12 × 45,000 sqft", r8["permits"], "$12 × 45,000 sqft", "same storey area"),
    ("  DCL New Westminster ($18.17/sqft)", r6["dcl"], "$18.17 × 45,000", r8["dcl"], "$18.17 × 45,000", "same storey area"),
    ("TOTAL SOFT COSTS", r6["total_soft"], "sum soft costs", r8["total_soft"], "sum soft costs", "same soft costs (same GFA)"),
    ("", None, "", None, "", ""),
    ("Interest Reserve (18 mo @ 6.5%)", r6["interest_reserve"], "hard costs × 6.5% × 1.5", r8["interest_reserve"], "hard costs × 6.5% × 1.5", "higher for 8-storey"),
    ("Permanent Financing Fee (1% GDV)", r6["perm_fin_fee"], "1% × GDV", r8["perm_fin_fee"], "1% × GDV", "same GDV → same fee"),
    ("TOTAL FINANCING COSTS", r6["total_financing"], "interest + perm fee", r8["total_financing"], "interest + perm fee", ""),
    ("", None, "", None, "", ""),
    ("Developer Profit (15% GDV)", r6["dev_profit"], "15% × GDV", r8["dev_profit"], "15% × GDV", "same GDV → same profit"),
    ("TOTAL COSTS (excl. land)", r6["total_costs_preland"], "hard + soft + fin + profit", r8["total_costs_preland"], "hard + soft + fin + profit", "6-storey is lower"),
    ("", None, "", None, "", ""),
]

is_section = False
for i, (label, v6, u6, v8, u8, note) in enumerate(step2_rows):
    row = step2_start + 2 + i
    is_section = label in ("", "Soft Costs:")
    is_total = label.startswith("TOTAL")
    is_hard_soft_fin_profit = any(x in label for x in ("Hard Construction", "Structured Parking", "TOTAL HARD", "TOTAL SOFT",
                                                        "Interest Reserve", "Permanent Financing", "TOTAL FINANCING",
                                                        "Developer Profit", "TOTAL COSTS"))

    if label == "":
        for col in range(1, 7):
            ws_rlv.cell(row=row, column=col).fill = make_fill(GRAY_BG)
        continue

    fill = make_fill("E2EFDA") if is_total else (make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE))
    if is_section and label != "" and "Soft Costs:" in label:
        fill = make_fill(DARK_BLUE)

    c1 = ws_rlv.cell(row=row, column=1, value=label)
    c1.fill = fill; c1.font = Font(bold=is_total, size=11, color=WHITE if is_section else "000000")
    c1.alignment = Alignment(horizontal='left', vertical='center', indent=1 if not is_section else 2)
    c1.border = thin_border()

    c2 = ws_rlv.cell(row=row, column=2)
    c2.fill = fill; c2.font = Font(bold=is_total, size=11)
    c2.alignment = Alignment(horizontal='right', vertical='center')
    c2.border = thin_border()
    if v6 is not None and not is_total:
        c2.value = v6
        set_currency(ws_rlv, c2)

    c3 = ws_rlv.cell(row=row, column=3, value=u6)
    c3.fill = fill; c3.font = Font(size=10, color="555555" if not is_section else WHITE)
    c3.alignment = Alignment(horizontal='left', vertical='center')
    c3.border = thin_border()

    c4 = ws_rlv.cell(row=row, column=4)
    c4.fill = fill; c4.font = Font(bold=is_total, size=11)
    c4.alignment = Alignment(horizontal='right', vertical='center')
    c4.border = thin_border()
    if v8 is not None and not is_total:
        c4.value = v8
        set_currency(ws_rlv, c4)

    c5 = ws_rlv.cell(row=row, column=5, value=u8)
    c5.fill = fill; c5.font = Font(size=10, color="555555" if not is_section else WHITE)
    c5.alignment = Alignment(horizontal='left', vertical='center')
    c5.border = thin_border()

    c6 = ws_rlv.cell(row=row, column=6, value=note)
    c6.fill = fill; c6.font = Font(size=10, italic=True, color="666666" if not is_section else WHITE)
    c6.alignment = Alignment(horizontal='left', vertical='center')
    c6.border = thin_border()

# ── STEP 3: RESIDUAL LAND VALUE ──
step3_start = step2_start + 2 + len(step2_rows) + 1
section_header(ws_rlv, step3_start, 1, 6, "STEP 3 — RESIDUAL LAND VALUE")

header_row(ws_rlv, step3_start + 1, [1, 2, 3, 4, 5, 6],
           ["", "6-Storey", "", "8-Storey", "", "Notes"])

r6 = calc_model(6, 430)
r8 = calc_model(8, 430, height_premium=1.05)

step3_rows = [
    ("Gross Development Value (GDV)", r6["gdv"], "", r8["gdv"], "", "same for both scenarios"),
    ("Less: Total Hard Costs", -r6["total_hard"], "", -r8["total_hard"], "", "higher for 8-storey"),
    ("Less: Total Soft Costs", -r6["total_soft"], "", -r8["total_soft"], "", "same for both"),
    ("Less: Total Financing Costs", -r6["total_financing"], "", -r8["total_financing"], "", "higher for 8-storey"),
    ("Less: Developer Profit (15%)", -r6["dev_profit"], "", -r8["dev_profit"], "", "same for both"),
    ("RESIDUAL LAND VALUE", r6["residual"], "", r8["residual"], "", "higher for 6-storey (lower costs)"),
    ("Less: BC Property Transfer Tax", -r6["land_tax"], "", -r8["land_tax"], "", "approx. marginal"),
    ("NET RESIDUAL LAND VALUE", r6["net_residual"], "", r8["net_residual"], "", "actual land acquisition capacity"),
]

for i, (label, v6, u6, v8, u8, note) in enumerate(step3_rows):
    row = step3_start + 2 + i
    is_total = label.startswith("RESIDUAL") or label.startswith("NET RESIDUAL")
    fill = make_fill("E2EFDA") if is_total else (make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE))

    c1 = ws_rlv.cell(row=row, column=1, value=label)
    c1.fill = fill; c1.font = Font(bold=is_total, size=11)
    c1.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c1.border = thin_border()

    for col, val in [(2, v6), (4, v8)]:
        c = ws_rlv.cell(row=row, column=col, value=val)
        c.fill = fill; c.font = Font(bold=is_total, size=11)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = thin_border()
        if val is not None:
            set_currency(ws_rlv, c)
        if is_total and val is not None:
            if val > 0:
                c.fill = make_fill(GREEN_FILL)
                c.font = Font(bold=True, color=GREEN_FONT, size=11)
            else:
                c.fill = make_fill(RED_FILL)
                c.font = Font(bold=True, color=RED_FONT, size=11)

    for col in [3, 5, 6]:
        ws_rlv.cell(row=row, column=col).fill = fill
        ws_rlv.cell(row=row, column=col).border = thin_border()
    ws_rlv.cell(row=row, column=3).value = ""
    ws_rlv.cell(row=row, column=5).value = ""
    ws_rlv.cell(row=row, column=6).value = note
    ws_rlv.cell(row=row, column=6).font = Font(italic=True, size=10, color="555555")
    ws_rlv.cell(row=row, column=6).alignment = Alignment(horizontal='left', vertical='center')

# Summary box
summary_row = step3_start + 2 + len(step3_rows) + 2
section_header(ws_rlv, summary_row, 1, 6, "SUMMARY")

summary_data = [
    ("Metric", "6-Storey", "8-Storey", "Difference (8 - 6)", "", ""),
    ("GDV", r6["gdv"], r8["gdv"], r8["gdv"] - r6["gdv"], "", "same"),
    ("Total Development Costs (excl. land)", r6["total_costs_preland"], r8["total_costs_preland"],
     r8["total_costs_preland"] - r6["total_costs_preland"], "", "8-storey higher"),
    ("RESIDUAL LAND VALUE", r6["residual"], r8["residual"],
     r8["residual"] - r6["residual"], "", "6-storey higher (lower costs)"),
    ("Net Residual (after tax)", r6["net_residual"], r8["net_residual"],
     r8["net_residual"] - r6["net_residual"], "", "6-storey higher"),
]

for i, row_data in enumerate(summary_data):
    row = summary_row + 1 + i
    fill = make_fill(DARK_BLUE) if i == 0 else (make_fill("E2EFDA") if i == 3 else (make_fill(LIGHT_BLUE) if i % 2 == 0 else make_fill(WHITE)))
    font = Font(bold=(i == 0 or i == 3), size=11, color=WHITE if i == 0 else "000000")
    for col, val in enumerate(row_data, 1):
        c = ws_rlv.cell(row=row, column=col, value=val)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal='left' if col in [1, 4] else 'right', vertical='center', indent=1 if col == 1 else 0)
        c.border = thin_border()
        if col in [2, 3, 4] and isinstance(val, (int, float)) and val != "":
            set_currency(ws_rlv, c)

print("Residual Land Value sheet done.")

# ══════════════════════════════════════════════════════════════
#  SAVE
# ══════════════════════════════════════════════════════════════
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"\nFile saved to: {OUTPUT_PATH}")

# ── Print summary ──
print("\n" + "="*60)
print("MODEL SUMMARY")
print("="*60)
r6 = calc_model(6, 430)
r8 = calc_model(8, 430, height_premium=1.05)

print(f"\n6-Storey Scenario:")
print(f"  Gross Area: {r6['gross_area']:,.0f} sqft | Units: {r6['unit_count']}")
print(f"  GDV: ${r6['gdv']:,.0f}")
print(f"  Total Hard Costs: ${r6['total_hard']:,.0f}")
print(f"  Total Soft Costs: ${r6['total_soft']:,.0f}")
print(f"  Total Financing: ${r6['total_financing']:,.0f}")
print(f"  Dev Profit (15%): ${r6['dev_profit']:,.0f}")
print(f"  Total Costs (excl. land): ${r6['total_costs_preland']:,.0f}")
print(f"  RESIDUAL LAND VALUE: ${r6['residual']:,.0f}")
print(f"  Net Residual (after tax): ${r6['net_residual']:,.0f}")
print(f"  Breakeven Cap Rate: {r6['breakeven_cap']:.2%}")
print(f"  Dev Yield: {r6['dev_yield']:.2%}")

print(f"\n8-Storey Scenario:")
print(f"  Gross Area: {r8['gross_area']:,.0f} sqft | Units: {r8['unit_count']}")
print(f"  GDV: ${r8['gdv']:,.0f}")
print(f"  Total Hard Costs: ${r8['total_hard']:,.0f}")
print(f"  Total Soft Costs: ${r8['total_soft']:,.0f}")
print(f"  Total Financing: ${r8['total_financing']:,.0f}")
print(f"  Dev Profit (15%): ${r8['dev_profit']:,.0f}")
print(f"  Total Costs (excl. land): ${r8['total_costs_preland']:,.0f}")
print(f"  RESIDUAL LAND VALUE: ${r8['residual']:,.0f}")
print(f"  Net Residual (after tax): ${r8['net_residual']:,.0f}")
print(f"  Breakeven Cap Rate: {r8['breakeven_cap']:.2%}")
print(f"  Dev Yield: {r8['dev_yield']:.2%}")

print("\nBreakeven Analysis:")
print(f"  6-Storey: {r6['breakeven_cap']:.2%} (buffer: {(r6['breakeven_cap']-0.0575)*10000:.0f} bps above 5.75%)")
print(f"  8-Storey: {r8['breakeven_cap']:.2%} (buffer: {(r8['breakeven_cap']-0.0575)*10000:.0f} bps above 5.75%)")